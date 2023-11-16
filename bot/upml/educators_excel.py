import functools

import asyncio
import datetime as dt
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import cell
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from bot.database import database_init
from bot.database.repository.repository import Repository
from bot.settings import get_settings
from bot.utils.datehelp import date_today, datetime_now


if TYPE_CHECKING:
    from bot.database.repository import EducatorsScheduleRepository


async def load_educators_from_xlsx(repo: "EducatorsScheduleRepository") -> None:

    today = date_today()
    current_date = dt.date(year=today.year, month=today.month, day=1)
    while current_date.month == today.month:
        educators_names = parse_excel(current_date.day)
        schedule = format_text(educators_names)

        await repo.save_or_update_to_db(
            date=current_date,
            schedule_text=schedule,
        )

        current_date += dt.timedelta(days=1)

def parse_excel(day: int) -> list[str]:
    wb: Workbook = load_workbook(Path().cwd() / "resources" / "educators.xlsx")
    wb_room: Workbook = load_workbook(Path().cwd() / "resources" / "educators_room.xlsx")
    ws: Worksheet = wb.active
    ws_room: Worksheet = wb_room.active

    list_educators: list[Educator] = []
    index_educators: int = 3
    work_cell: cell = ws.cell(index_educators, day + 2)

    while not work_cell.value is None:
        if "В" not in work_cell.value:
            educator: Educator = Educator(day, index_educators, ws, ws_room)
            list_educators.append(educator)
        index_educators += 1
        work_cell = ws.cell(index_educators, day + 2)

    list_educators.sort(key=educators2time)
    today = date_today()
    day = dt.date(year=today.year, month=today.month, day=day) - dt.timedelta(1)
    day = day.day
    index_educators = 3
    work_cell = ws.cell(index_educators, day + 2)

    while not work_cell.value is None:
        if "В" not in work_cell.value:
            if work_cell.value.split("-")[1][0] == "9":
                educator: Educator = Educator(day, index_educators, ws, ws_room)
                list_educators.insert(0, educator)
                list_educators[0].barder_start_time()
        index_educators += 1
        work_cell = ws.cell(index_educators, day + 2)

    isrooms: bool = list_educators[0].isroom and list_educators[1].isroom
    if isrooms and list_educators[0].room > list_educators[1].room:
        list_educators[0], list_educators[1] = list_educators[1], list_educators[0]
    isrooms: bool = list_educators[-1].isroom and list_educators[-2].isroom
    if isrooms and list_educators[-2].room > list_educators[-1].room:
        list_educators[-2], list_educators[-1] = list_educators[-1], list_educators[-2]

    #map(functools.partial(Educator.compose_string, isname=True, isnickname=True), list_educators)

    list_text_educators: list[str] = map(Educator.compose_string, list_educators)
    return list_text_educators

class Educator:
    def __init__(self, day: int, index_educators: int, ws: Worksheet, ws_room: Worksheet):
        time = str(ws.cell(index_educators, day + 2).value).split("-")
        self.time_start: str = time[0]
        self.time_end: str = time[1]
        self.name: str = ws.cell(index_educators, 1).value
        self.nickname: str = ws.cell(index_educators, 2).value
        self.isroom: bool = not ws_room.cell(index_educators - 1, day + 1).value is None
        if self.isroom:
            self.room: int = int(ws_room.cell(index_educators - 1, day + 1).value)

    def compose_string(self, isname: bool = True, isnickname: bool = False) -> str:
        self.format_time()
        line: str = self.time_start + "-" + self.time_end + " "
        if isnickname and isname:
            line += self.name + " (" + self.nickname + ")"
        elif isnickname:
            line += self.nickname
        else:
            line += self.name
        if self.isroom:
            line += " ("
            line += "3-6" if self.room == 59 else "7-9"
            line += " этаж)"
        return line
    def barder_start_time(self) -> None:
        self.time_start = "00"
    def format_time(self) -> str:
        if int(self.time_start) < int(self.time_end):
            self.time_start = self.time_start + ":00"
            self.time_end = self.time_end + ":00"
        else:
            self.time_start = self.time_start + ":00"
            self.time_end = "23:59"


def _parse_excel(day: int) -> list[str]:
    wb: Workbook = load_workbook(Path().cwd() / "resources" / "educators.xlsx")
    ws: Worksheet = wb.active

    list_educators: list[str] = []
    index_educarors: int = 3
    work_cell: cell = ws.cell(index_educarors, day + 2)

    while not work_cell.value is None:
        if "В" not in work_cell.value:
            line: str = work_cell.value + " " + ws.cell(index_educarors, 1).value
            list_educators.append(line)
        index_educarors += 1
        work_cell = ws.cell(index_educarors, day + 2)

    list_educators.sort(key=educators2time)
    today = date_today()
    day = dt.date(year=today.year, month=today.month, day=day).replace(1).day
    index_educarors = 3
    work_cell: cell = ws.cell(index_educarors, day + 2)
    list_educators[-1] = trimming_time_from_end(list_educators[-1], "23:59")
    list_educators[-2] = trimming_time_from_end(list_educators[-2], "23:59")

    while not work_cell.value is None:
        if "В" not in work_cell.value:
            if work_cell.value.split("-")[1][0] == "9":
                line: str = work_cell.value + " " + ws.cell(index_educarors, 1).value
                list_educators.insert(0, line)
                list_educators[0] = trimming_time_from_start(list_educators[0], "00:00")
        index_educarors += 1
        work_cell = ws.cell(index_educarors, day + 2)

    list_educators = map(functools.partial(Educator.compose_string, isname=False, isnickname=True), list_educators)
    return list_educators
def trimming_time_from_end(line: str, time: str) -> str:
    list_line: list[str] = line.split()
    list_time: list[str] = [list_line[0].split("-")[0], list_line[0].split("-")[1]]
    list_name: list[str] = [list_line[1], list_line[2], list_line[3]]
    line: str = list_time[0] + "-" + time + " " + " ".join(list_name)
    return line

def trimming_time_from_start(line: str, time: str) -> str:
    list_line: list[str] = line.split()
    list_time: list[str] = [list_line[0].split("-")[0], list_line[0].split("-")[1]]
    list_name: list[str] = [list_line[1], list_line[2], list_line[3]]
    line: str = time + "-" + list_time[1] + " " + " ".join(list_name)
    return line

def format_educators_line(line: str) -> str:
    list_line: list[str] = line.split()
    list_time: list[str] = [list_line[0].split("-")[0], list_line[0].split("-")[1]]
    list_name: list[str] = [list_line[1], list_line[2], list_line[3]]
    list_time[0] += ":00" if ":" not in list_time[0] else ""
    list_time[1] += ":00" if ":" not in list_time[1] else ""
    line: str = list_time[0] + "-" + list_time[1] + " " + " ".join(list_name)
    return line
def educators2time(educators: Educator) -> int:
    return int(educators.time_start)
def format_text(educators_names: list[str]) -> str:
    return '\n'.join(educators_names)


async def main() -> None:
    settings = get_settings()
    session_maker: "async_sessionmaker[AsyncSession]" = await database_init(settings.db)

    async with session_maker.begin() as session:
        repo = Repository(session)
        educators_repo = repo.educators
        await load_educators_from_xlsx(educators_repo)


if __name__ == "__main__":
    asyncio.run(main())
