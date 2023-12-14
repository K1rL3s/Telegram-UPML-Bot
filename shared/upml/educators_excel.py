import asyncio
import datetime as dt
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from shared.core.settings import get_settings
from shared.database import database_init
from shared.database.repository.repository import Repository
from shared.utils.datehelp import date_today

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from shared.database.repository import EducatorsScheduleRepository


class FormatNameFlags(BaseModel):
    is_surname: bool = False
    is_name: bool = True
    is_patronymic: bool = True
    is_nickname: bool = False


async def load_educators_from_xlsx(repo: "EducatorsScheduleRepository") -> None:
    today = date_today()
    current_date = dt.date(year=today.year, month=today.month, day=1)
    resoures_path = (Path().cwd() / "shared" / "resources").resolve()

    wb: "Workbook" = load_workbook(resoures_path / "educators.xlsx")
    wb_room: "Workbook" = load_workbook(resoures_path / "educators_room.xlsx")
    schedule_ws: "Worksheet" = wb.active
    rooms_ws: "Worksheet" = wb_room.active

    while current_date.month == today.month:
        if await repo.get(current_date) is None:
            schedule = parse_educators_excels(current_date, schedule_ws, rooms_ws)
            await repo.save_or_update_to_db(
                date=current_date,
                schedule_text=schedule,
            )
        current_date += dt.timedelta(days=1)


def parse_educators_excels(
    day: "dt.date",
    schedule_ws: "Worksheet",
    schedule_rooms_ws: "Worksheet",
) -> str:
    yesterday = day - dt.timedelta(1)

    today_educators = EducatorsList(
        day,
        schedule_ws,
        schedule_rooms_ws,
    )
    today_educators.educators().sort(key=Educator.time_start)
    yesterday_educators = EducatorsYesterdayNightList(
        yesterday,
        schedule_ws,
        schedule_rooms_ws,
    )

    educators = yesterday_educators.educators() + today_educators.educators()
    EducatorsList.swap_educators_by_floors(educators)

    schedule: list[str] = [
        educator.compose_string(FormatNameFlags()) for educator in educators
    ]
    return "\n".join(schedule)


def is_weekend(cell: Cell) -> bool:
    return "В" in cell.value


def is_night_shift(cell: Cell) -> bool:
    return cell.value.split("-")[1][0] == "9"


class Educator:
    def __init__(
        self,
        day: int,
        index_educators: int,
        schedule_ws: "Worksheet",
        rooms_ws: "Worksheet",
    ):
        time = str(schedule_ws.cell(index_educators, day + 2).value).split("-")
        full_name: list[str] = schedule_ws.cell(index_educators, 1).value.split()

        self._time_start: int = int(time[0])
        self._time_end: int = int(time[1])
        self._surname: str = full_name[0]
        self._name: str = full_name[1]
        self._patronymic: str = full_name[2]
        self._nickname: str = schedule_ws.cell(index_educators, 2).value
        self._is_room: bool = rooms_ws.cell(index_educators, day + 2).value is not None
        self._room: int | None = None
        if self._is_room:
            self._room: int = int(rooms_ws.cell(index_educators, day + 2).value)

    def compose_string(
        self,
        format_name_flags: FormatNameFlags,
    ) -> str:
        start_time, end_time = self._format_time()
        full_name = self._format_name(format_name_flags)
        room = self._format_room()
        return f"<b>{start_time}-{end_time}</b> {full_name} {room}"

    def _format_name(
        self,
        format_name_flags: FormatNameFlags,
    ) -> str:
        full_name = []
        if format_name_flags.is_surname:
            full_name.append(self._surname)
        if format_name_flags.is_name:
            full_name.append(self._name)
        if format_name_flags.is_patronymic:
            full_name.append(self._patronymic)

        if format_name_flags.is_nickname:
            if full_name:
                full_name.append(f"({self._nickname})")
            else:
                full_name.append(self._nickname)

        return " ".join(full_name)

    def _format_room(self) -> str:
        if self._room == 59:
            return "(3-6 этаж)"
        if self._room == 79:
            return "(7-9 этаж)"
        return ""

    def zero_start_time(self) -> None:
        self._time_start = 0

    def _format_time(self) -> tuple[str, str]:
        if self._time_start < self._time_end:
            return f"{self._time_start}:00", f"{self._time_end}:00"
        return f"{self._time_start}:00", "23:59"

    def time_start(self) -> int:
        return self._time_start

    @property
    def is_room(self) -> bool:
        return self._is_room

    @property
    def room(self) -> int | None:
        return self._room


class EducatorsList:
    def __init__(
        self, day: "dt.date", schedule_ws: "Worksheet", rooms_ws: "Worksheet"
    ) -> None:
        self._schedule_ws = schedule_ws
        self._rooms_ws = rooms_ws
        self._educators: list["Educator"] = []
        self._filter_educators_by_date(day)

    def _filter_educators_by_date(self, day: "dt.date") -> None:
        index_educators = 2
        day_index = day.day + 2
        current_cell: Cell = self._schedule_ws.cell(index_educators, day_index)

        while current_cell.value is not None:
            if not is_weekend(current_cell):
                educator = Educator(
                    day.day,
                    index_educators,
                    self._schedule_ws,
                    self._rooms_ws,
                )
                self._educators.append(educator)
            index_educators += 1
            current_cell = self._schedule_ws.cell(index_educators, day_index)

    @staticmethod
    def swap_educators_by_floors(educators: list["Educator"]) -> None:
        if not len(educators) > 1:
            return

        is_rooms = educators[0].is_room and educators[1].is_room
        is_wrong_order = is_rooms and educators[0].room > educators[1].room
        if is_wrong_order:
            educators[0], educators[1] = (
                educators[1],
                educators[0],
            )

        is_rooms = educators[-1].is_room and educators[-2].is_room
        is_wrong_order = is_rooms and educators[-2].room > educators[-1].room
        if is_wrong_order:
            educators[-2], educators[-1] = (
                educators[-1],
                educators[-2],
            )

    def educators(self) -> list["Educator"]:
        return self._educators


class EducatorsYesterdayNightList(EducatorsList):
    def _filter_educators_by_date(self, day: "dt.date") -> None:
        index_educators = 2
        day_index = day.day + 2
        current_cell: Cell = self._schedule_ws.cell(index_educators, day_index)

        while current_cell.value is not None:
            if not is_weekend(current_cell) and is_night_shift(current_cell):
                educator = Educator(
                    day.day,
                    index_educators,
                    self._schedule_ws,
                    self._rooms_ws,
                )
                self._educators.insert(0, educator)
                self._educators[0].zero_start_time()
            index_educators += 1
            current_cell = self._schedule_ws.cell(index_educators, day_index)


async def main() -> None:
    settings = get_settings()
    session_maker: "async_sessionmaker[AsyncSession]" = await database_init(settings.db)

    async with session_maker.begin() as session:
        repo = Repository(session)
        educators_repo = repo.educators
        await load_educators_from_xlsx(educators_repo)


if __name__ == "__main__":
    asyncio.run(main())
